# -*- coding: utf-8 -*-
"""
Created on Thu Apr 16 19:48:35 2020

@author: maxaj
"""
import numpy as np
import matplotlib.pyplot as PL
import CoolProp.CoolProp as CP
from openpyxl import *
import math


n = np.linspace(1,365,365)

delta_list = []
delta2_list = []

for i in range(len(n)):

    #Calculate B (in degrees) for use in the delta calculation
    B_deg = (n[i]-1)*(360/365)
    
    #Convert B into radians
    B = B_deg*np.pi/180
    
    #Calculate delta (in degrees) - declination - angle of the sun at solar noon i.e. whent the sun is on the local meridian
    delta_deg = (180/(np.pi))*(0.006918-0.399912*np.cos(B)+0.070257*np.sin(B)-0.006758*np.cos(2*B)+0.000907*np.sin(2*B)-0.002697*np.cos(3*B)+0.00148*np.sin(3*B))
    
    #Convert delta into radians
    delta = delta_deg*np.pi/180

    delta_2 = 23.45*np.sin((360*((284+n[i])/365))*np.pi/180)
    
    delta_list.append(delta_deg)
    delta2_list.append(delta_2)

print(n)
print(delta_list)
print(delta2_list)

wb = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Monthly_Collector_Output_Data_2.xlsx")

#Define which worksheet to write data to
ws = wb["Error_Analysis"] 

Month_list = ['January','February','March','April','May','June','July','August','September','October','November','December']

for i in range(len(delta_list)):

    #Define cell to write data to
    wcell1 = ws.cell(i+3,110)
    wcell2 = ws.cell(i+3,111)
    wcell3 = ws.cell(i+3,112)
    #wcell4 = ws.cell(i+3,32)
    #wcell5 = ws.cell(i+2,53)

    #State what value should be written to the cell
    wcell1.value = n[i]
    wcell2.value = delta_list[i]
    wcell3.value = delta2_list[i]
    #wcell4.value = Month_list[i]
    #wcell5.value = Tpm_array[i]

#Save the spreadsheet
wb.save("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Monthly_Collector_Output_Data_2.xlsx")