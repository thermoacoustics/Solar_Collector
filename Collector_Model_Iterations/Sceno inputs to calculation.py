# -*- coding: utf-8 -*-
"""
Created on Mon Mar  9 22:00:15 2020

@author: maxaj
"""

import numpy as np
import matplotlib.pyplot as PL
import CoolProp.CoolProp as CP
from openpyxl import *

#Load workbook to load data from
wb1 = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Athens_Climate_Data.xlsx")

#Define which worksheet to read data from
ws1 = wb1["Sheet1"]

#Define cell to read data from
wcell1 = ws1.cell(6,1).value 

I = ws1.cell(16,3).value #Read total radiation incident on a flat surface horizonatal in Watts per metre squared from Sceno weather data spreadsheet
Ibn = ws1.cell(16,4).value #Input beam radiation incident on a flat surface perpendicular to the beam radiation in Watts per metre squared from Sceno weather data spreadsheet
Pg = 0.5 #Input ground reflected radiation factor
beta_deg = 25 #Input collector slope in degrees
beta = beta_deg*np.pi/180 #Convert collector slope into radians
lat_deg = 38 #Input lattitude of the collector location in degrees
lat = lat_deg*np.pi/180 #Convert collector lattitude into radians
n = 1 #Input day of the year (1st January = 1)
h = ws1.cell(16,1).value #Input hour of the day (00:00-01:00 = 1)
Ula = 0.504 #Input the part of the total loss coefficient that doesn't depend on temperature in Watts per metres squared kelvin
Ulb = 0.006 #Input the temperature dependent part of the total loss coefficient in watts per metres squared kelvin
t = 0.926 #Input collector transmittance
a = 0.95 #Input collector absorbance
Ag = 1.96 #Input collector gross area in metres squared
Ac = 1.84 #Input collector aperture area in metres squared
Ta = ws1.cell(16,2).value #Read the ambient temperature in celsius from the Sceno weather spreadsheet
m_dot = 0.0392 #Input the mass flow rate to the collector in kg per second
F_dash = 0.968 #Input collector efficiency factor
P = 16*(10**5) #Input collector fluid pressure in bar


#Create list of fluid inlet temperature values to be tested in celsius
#Tfi_list = [40.0,50.0,60.0,70.0,80.0,90.0,100.0,110.0,120.0,130.0,140.0,150.0,160.0,170.0,180.0,190.0,200.0,220.0]
#Tfi_list = np.linspace(40,220,100)

#Convert the list into an array
#Tfi_array = np.array(Tfi_list)

#Create list of the 3 mean fluid temperature to be used in celsius
#Tfm_list = (25,50,75)

#Convert the list into an array
#Tfm_array = np.array(Tfm_list)

#print(len(Tfi_list))

Tfi = 40 #Input initial fluid inlet temperature guess in celsius
Tfm_tar = 50 #Input the target mean fluid temperature
    
def absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar):
    
    #Calculate total radiation incident on a flat surface
    #I = Ib+Id
    #print('I:',I)
    
    #Calculate B (in degrees) for use in the delta calculation
    B_deg = (n-1)*(360/365)
    
    #Convert B into radians
    B = B_deg*np.pi/180
    
    #Calculate delta (in degrees) - declination - angle of the sun at solar noon i.e. whent the sun is on the local meridian
    delta_deg = (180/(np.pi))*(0.006918-0.399912*np.cos(B)+0.070257*np.sin(B)-0.006758*np.cos(2*B)+0.000907*np.sin(2*B)-0.002697*np.cos(3*B)+0.00148*np.sin(3*B))
    
    #Convert delta into radians
    delta = delta_deg*np.pi/180
    
    #Calculate omega (in degrees) - hour angle
    omega_deg = h*15-187.5
    #Convert omega to radians
    omega = omega_deg*np.pi/180
    
    #Calculate theta (in radians) - Angle of incidence on the collector plate
    theta = np.arccos((np.sin(delta)*np.sin(lat-beta))+(np.cos(delta)*np.cos(omega)*np.cos(lat-beta)))
    
    #Convert theta to degrees for use in angular beam transmition absorption product calculation
    theta_deg = theta*180/np.pi
    
    #Calculate the beam radiation incident on the tilted collector surface
    Ibt = Ibn*(np.cos(theta))

    #Calculate the beam radiation incident on a horizontal surface
    Ibh = Ibn*((np.sin(delta))*(np.sin(lat))+(np.cos(delta))*(np.cos(omega))*(np.cos(lat)))
    
    #Calculate the diffuse radiation incident on a horizontal surface
    Idh = I - Ibh
    
    #Calculate equivalent theta for ground reflected radiation
    theta_tag = 90-(0.5788*beta)+(0.002693*(beta**2))
    #print('theta_tag:',theta_tag)
    
    #Calculate equivalent theta for diffuse radiation
    theta_tad = 59.7-(0.1388*beta)+(0.001497*(beta**2))
    #print('theta_tad:',theta_tad)
    
    #Calculate normal transimition absorption product
    tan = 1.01*t*a
    #print('tan:', tan)
    
    #Calculate the angular transmition absorption product for beam radiation - using IAM specific to TVP collector
    if theta_deg <= 30:
        #If theta is less than 30 then transmission absorption product is just equal to the normal value
        tab = tan
    elif theta_deg > 30 and theta_deg < 90:
        #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
        tab = tan*((-1*(10**(-8))*theta_deg**4)+(-2*(10**(-6))*theta_deg**3)+(0.0002*theta_deg**2)-(0.003*theta_deg)+1.0061)
    else:
        tab = 0
        print('No beam radiation incident on the collector')
        
    #print('tab:',tab)
    
    #Calculate the angular transmition absorption product for ground reflected radiation
    if theta_tag <= 30:
        #If theta is less than 30 then transmission absorption product is just equal to the normal value
        tag = tan
    elif theta_tag > 30 and theta_tag < 90:
        #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
        tag = tan*((-1*(10**(-8))*theta_tag**4)+(-2*(10**(-6))*theta_tag**3)+(0.0002*theta_tag**2)-(0.003*theta_tag)+1.0061)
    else:
        tag = 0
        
    #print('tag:',tag)
    
    #Calculate the angular transmition absorption product for diffuse radiation
    if theta_tad <= 30:
        #If theta is less than 30 then transmission absorption product is just equal to the normal value
        tad = tan
    elif theta_tad > 30 and theta_tad < 90:
        #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
        tad = tan*((-1*(10**(-8))*theta_tad**4)+(-2*(10**(-6))*theta_tad**3)+(0.0002*theta_tad**2)-(0.003*theta_tad)+1.0061)
    else:
        tad = 0
        
    #print('tad:', tad)
    
    #Calculate Rb
    #Rb = ((np.cos(lat-beta))*(np.cos(delta))*(np.cos(omega))+(np.sin(lat-beta))*(np.sin(delta)))/((np.cos(lat))*(np.cos(delta))*(np.cos(omega))+(np.sin(lat))*(np.sin(delta)))
    #Rb = 1
    
    #Calculate the absorbed solar radiation per unit collector area
    #S = Ib*Rb*tab + Id*tad*((1+np.cos(beta))/2) + I*Pg*tag*((1-np.cos(beta))/2)
    #print('S:',S)
    
    #Calculate the absorbed solar radiation per unit collector area
    S = Ibt*tab + Idh*tad*((1+np.cos(beta))/2) + I*Pg*tag*((1-np.cos(beta))/2)
    #print('S:',S)
    
    #Include factor for reduced absorption due to dust
    Df = 0.99

    #Include factor for reduced absoprtion due to self shading
    Ds = 0.99
    
    #Calculated adjusted shading when dust and self shading are taken into account
    Sa = S*Df*Ds
    #print('Sa:', Sa)
    
    #Calaculate initial Cp estimate
    Cp = CP.PropsSI('Cp0mass','T',Tfi+273,'P',P,'water')
    
    #Calculate initial Ul estimate
    Ul = Ulb*(Tfi-Ta)+Ula
    
    #Calculate initial Fr estimate
    Fr = ((m_dot*Cp)/(Ac*Ul))*(1-np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp)))
    
    #Calculate initial estimate of the useful energy gain
    Qu = Ac*Fr*(Sa-Ul*(Tfi-Ta))
    
    #Set limit for the number of iterations that can be carried out
    Q = 150
    
    for r in range(Q):
        if r==(Q-2):
            print('Tfm iteration has not converged to the required value')
    
        #Define i so that a counter for the loop can be used
        i = 2
        
        #Set a limit for the number of iterations that can be carried out
        Z = 100
    
        for n in range(Z):
        
            #Use counter to inform user on whether or not the iteration has converged to within 1%
            i += 1
            if i == Z:
                print('Qu iteration has not converged to within 1%')
        
            #Calculate new Tfm value
            Tfm = Tfi+(Qu/(Ac*Fr*Ul))*(1-(Fr/F_dash))
            
            #Calculate new Ul value using Tfm
            Ul = (49/4600)*(Tfm-Ta)+(343/460)
            
            #Calculate new Cp value using Tfm
            Cp = CP.PropsSI('Cp0mass','T',Tfm+273,'P',P,'water')
        
            #Calculate new Fr value with new Cp value
            Fr = ((m_dot*Cp)/(Ac*Ul))*(1-np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp)))
        
            #Calculate Tpm estimate
            Tpm = Tfi+(Qu/(Ac*Fr*Ul))*(1-Fr)
        
            #Introduce Qu_est so the value of Qu in this iteration can be compared to the value in the last iteration.
            Qu_est = Qu
        
            #Calculate new Qu estimate
            Qu = Ac*(Sa-Ul*(Tpm-Ta))
        
            #Terminate loop if the new value of Qu is within 1% of the previous value
            #if np.all(np.logical_and((Qu > 0.99*Qu_est),(Qu < 1.01*Qu_est))):
            if Qu > 0.99*Qu_est and Qu < Qu_est:
                break
            else:
                continue
            
        #Provide if function so that an iteration to the desired mean temperature is carried out
        if Tfm > (Tfm_tar+0.1):
            Tfi -= 0.05
        elif Tfm < (Tfm_tar-0.1):
            Tfi += 0.05
        else:
            break
    
        #print('Tfm:',Tfm)
        #print('Cp:',Cp)
        #print('Fr:',Fr)
        #print('Tpm:',Tpm)
        #print('Qu:',Qu)
    
    #Set any results of negative power output to zero
    #for i in range(len(Qu)):
        #if Qu[i] < 0:
            #Qu[i]=0

    #Calculate the fluid outlet temperature
    Tfo = Ta + (Sa/Ul) + (Tfi-Ta-(Sa/Ul))*np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp))
    #print('Tfo',Tfo)
    
    #Calculate Collector Efficiency
    eff = Qu/(I*Ag)
    #print('Efficiency:',eff)
    
    #Calculate difference between the mean fluid temperature and ambient temperature
    T_diff = Tfm - Ta
    #print('T_diff:',T_diff)
    
    #Calculate difference between mean plate and mean fluid temperatures
    T_grad = Tpm - Tfm
    
    #Calculate beam contribution to radiation incident on the collector
    Ibt_col = Ac*Ibt
    
    #Calculate the diffuse contribution to the radiation incident on the collector
    Idt_col = Ac*Idh*((1+np.cos(beta))/2)
    
    #Calculate ground reflected contribution to the radiation incident on the collector
    Igt_col = Ac*I*Pg*((1-np.cos(beta))/2)
    
    #Calculate total radaition incident on the collector
    It_col = Ibt_col+Idt_col+Igt_col
    
    #Calculate the output energy for that hour
    E_out = Qu*3600
        
    #Put function outputs in a list
    Function_outputs = [Qu,Tfo,eff,T_diff,Ul,T_grad,Tfm,Tpm,It_col,Ibt_col,Idt_col,Igt_col,E_out]
    
    return Function_outputs

#Output an array of useful power results
Qu_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[0]
    
#Output an array of collector efficiency results
eff_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[2]

#Output an array of the results for the temperature difference between the mean fluid temperature and ambient temperature
T_diff_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[3]

#Output array for Ul results
Ul_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[4]

#Output array of plate to fluid temperature difference results
T_grad_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[5]

#Output array of fluid mean temperature
Tfm_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[6]

#Output array of plate mean temperature
Tpm_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[7]

#Output array of total incident radiation on the plate
It_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[8]

#Output array of total beam radiation on the plate
Ibt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[9]

#Output array of total diffuse radiation on the plate
Idt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[10]

#Output array of total diffuse radiation on the plate
Igt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[11]

#Output array for the energy output for that hour
E_out_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Tfm_tar)[12]
    
print('Tfm:',Tfm_array)
print('Qu:',Qu_array)
print('It:',It_col_array)
print('Ibt:',Ibt_col_array)
print('Idt:',Idt_col_array)
print('Igt:',Igt_col_array)
print('E_out:',E_out_array)

#print(absorbed_radiation(Ib,Id,Pg,beta,theta_deg,t,a,Ac,Ul,Tfi,Ta,m_dot,F_dash,P))
    
#PL.plot(T_diff_array,eff_array)
#PL.plot(T_diff_array,Ul_array)

#Create a workbook to write to
#wb = Workbook()
#wb = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Fifth attempt at writing to a spreadsheet.xlsx")

#Create a sheet to write to
#sheet1 = wb.add_sheet('Sheet 2')

#for i in range(len(eff_array)):
    
    #Write temperature difference data to the sheet
    #sheet1.write(i,2,T_diff_array[i])
    
    #Write efficiency data to the sheet
    #sheet1.write(i, 3, eff_array[i])

#Save the spreadsheet
#wb.save("Fifth attempt at writing to a spreadsheet.xls")

#Define which worksheet to write data to
#ws = wb["Sheet 2"] 

#for i in range(len(eff_array)):

    #Define cell to write data to
    #wcell1 = ws.cell(i+2,49)
    #wcell2 = ws.cell(i+2,50)
    #wcell3 = ws.cell(i+2,51)
    #wcell4 = ws.cell(i+2,52)
    #wcell5 = ws.cell(i+2,53)

    #State what value should be written to the 
    #wcell1.value = T_diff_array[i]
    #wcell2.value = eff_array[i]
    #wcell3.value = T_grad_array[i]
    #wcell4.value = Tfm_array[i]
    #wcell5.value = Tpm_array[i]

#Save the spreadsheet
#wb.save("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Fifth attempt at writing to a spreadsheet.xlsx")