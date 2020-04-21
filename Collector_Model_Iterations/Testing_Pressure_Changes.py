# -*- coding: utf-8 -*-
"""
Created on Sat Mar 21 21:29:18 2020

@author: maxaj
"""

import numpy as np
import matplotlib.pyplot as PL
import CoolProp.CoolProp as CP
from openpyxl import *

Tfi = 180 +273

P_list = np.linspace(1*(10**3),16*(10**5),40)

#P_list = []

Cp_list = []

for i in range(np.size(P_list)):
    
    P = P_list[i]

    Cp = CP.PropsSI('Cp0mass','T',Tfi,'P',P,'water')
    
    Cp_list.append(Cp)

print(Cp_list)

wb = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Full_Scenocalc_Report_Data_Comparison.xlsx")

#Create a sheet to write to
#sheet1 = wb.add_sheet('Sheet 2')

#for i in range(len(eff_array)):
    
    #Write temperature difference data to the sheet
    #sheet1.write(i,2,T_diff_array[i])
    
    #Write efficiency data to the sheet
    #sheet1.write(i, 3, eff_array[i])

#Save the spreadsheet
#wb.save("Fifth attempt at writing to a spreadsheet.xls")

#Define which worksheet to write data to
ws = wb["Pressure_Check"] 

for i in range(len(Cp_list)):

    #Define cell to write data to
    wcell1 = ws.cell(i+2,11)
    wcell2 = ws.cell(i+2,12)
    #wcell3 = ws.cell(i+2,4)
    #wcell4 = ws.cell(i+3,48)
    #wcell5 = ws.cell(i+2,53)

    #State what value should be written to the cell
    wcell1.value = P_list[i]
    wcell2.value = Cp_list[i]
    #wcell3.value = Tfm_list[i]
    #wcell4.value = T_diff_list[i]
    #wcell5.value = Tpm_array[i]

#Save the spreadsheet
wb.save("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Full_Scenocalc_Report_Data_Comparison.xlsx")