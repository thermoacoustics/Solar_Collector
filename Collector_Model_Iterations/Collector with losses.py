# -*- coding: utf-8 -*-
"""
Created on Thu Mar 19 16:59:51 2020

@author: maxaj
"""

import numpy as np
import matplotlib.pyplot as PL
import CoolProp.CoolProp as CP
from openpyxl import *
import math

#Load workbook to load data from
wb1 = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Full_Chad_Weather.xlsx")

#Define which worksheet to read data from
ws1 = wb1["Full_Chad_Weather"]

#I = ws1.cell(16,3).value #Read total radiation incident on a flat surface horizonatal in Watts per metre squared from Sceno weather data spreadsheet
#Ibn = ws1.cell(16,4).value #Input beam radiation incident on a flat surface perpendicular to the beam radiation in Watts per metre squared from Sceno weather data spreadsheet
Pg = 0.2 #Input ground reflected radiation factor
beta_deg = 13 #Input collector slope in degrees
beta = beta_deg*np.pi/180 #Convert collector slope into radians
lat_deg = 12.1 #Input lattitude of the collector location in degrees
lat = lat_deg*np.pi/180 #Convert collector lattitude into radians
#n = 1 #Input day of the year (1st January = 1)
#h = ws1.cell(16,1).value #Input hour of the day (00:00-01:00 = 1)
Ula = 0.504 #Input the part of the total loss coefficient that doesn't depend on temperature in Watts per metres squared kelvin
Ulb = 0.006 #Input the temperature dependent part of the total loss coefficient in watts per metres squared kelvin
t = 0.926 #Input collector transmittance
a = 0.95 #Input collector absorbance
Ag = 1.96 #Input collector gross area in metres squared
Ac = 1.84 #Input collector aperture area in metres squared
#Ta = ws1.cell(16,2).value #Read the ambient temperature in celsius from the Sceno weather spreadsheet
m_dot = 0.0392 #Input the mass flow rate to the collector in kg per second
F_dash = 0.968 #Input collector efficiency factor
P = 16*(10**5) #Input collector fluid pressure in bar
#Ud = 1 #Input pipe overall loss coefficient in watss per metres squared
Li = 3 #Input inlet pipe length in metres
Lo = 3 #Input outlet pipe length in metres
Di = 13*(10**(-3)) #Input inlet pipe diameter in metres
Do = 13*(10**(-3)) #Input outlet pipe diameter in metres


#Create list of fluid inlet temperature values to be tested in celsius
#Tfi_list = [40.0,50.0,60.0,70.0,80.0,90.0,100.0,110.0,120.0,130.0,140.0,150.0,160.0,170.0,180.0,190.0,200.0,220.0]
#Tfi_list = np.linspace(40,220,100)

#Convert the list into an array
#Tfi_array = np.array(Tfi_list)

#Create list of the 3 mean fluid temperature to be used in celsius
#Tfm_list = (25,50,75)

#Convert the list into an array
#Tfm_array = np.array(Tfm_list)

#print(len(Tfi_list))

Tfi = 100 #Input initial fluid inlet temperature guess in celsius
#Tfm_tar = 100 #Input the target mean fluid temperature

#Create list to put hourly horizontal radiation in
I_list = []

#Create list to put incident hourly radiation in
It_col_list = []

#Create list to put all the hourly energy output values in
Qu_list = []

#Create list to put the hourly mean fluid temperatures in
Tfm_list = []

#Create liste to put the hourly Tm - Ta values in
T_diff_list = []

#Create list to put hourly angles of incidence in
theta_deg_list = []

#Convert list into an array
#Qu_tot_array = np.array(Qu_tot_list)

#Define variable to be used to sum the hourly power output data
Qu_tot = 0

#Define variable to be used to sum the hourly incident radiation on the horizontal
I_sum = 0 

#Define variable to be used to sum the hourly incident radiation on the collector
It_col_sum = 0

#Define variable to be used to sum the hourly mean fluid temperatures
Tfm_sum = 0

#Define variable to be used to sum the hourly Tm - Ta values
T_diff_sum = 0

#Define list to append monthly energy outputs to
Qu_tot_list = []

#Define list to append monthly incident radiation values to
It_tot_list = []

#Define list to append the sum of all monthly mean fluid temperatures
Tfm_tot_list = []

#Define list to append the sum of all monthly Tm - Ta values
T_diff_tot_list = []

#Create a list of pipe loss coefficients to be tested
Ud_list = np.linspace(0.1,20,100)

#Convert collector slope list into an array
Ud_array = np.array(Ud_list)

#Loop over the required input values
for v in range(8760):
    I = ((ws1.cell(v+4,4).value)/3.6) #Read total radiation incident on a flat surface horizonatal in Watts per metre squared from Sceno weather data spreadsheet - convert from kJ/h to kWh
    Ibh = ((ws1.cell(v+4,6).value)/3.6) #Input beam radiation incident on a flat surface perpendicular to the beam radiation in Watts per metre squared from Sceno weather data spreadsheet - convert from kJ/h to kWh
    H = ws1.cell(v+4,1).value #Input hour of the day (00:00-01:00 = 1)
    Ta = ws1.cell(v+4,10).value #Read the ambient temperature in celsius from the Sceno weather spreadsheet
    n = math.ceil((H/24))
    h = H-(n-1)*24
    
    def absorbed_radiation(I,Ibh,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Ud_array,Li,Lo,Di,Do):
    
        #Calculate total radiation incident on a flat surface
        #I = Ib+Id
        #print('I:',I)
        
        #Allow function to be exited early if there is no solar radiation incident on the collector
        if I == 0 and Ibh == 0:
            Qu_loss = 0
            It_col = 0
            Function_outputs = [Qu_loss,It_col]
            return Function_outputs
        
        
        #Calculate B (in degrees) for use in the delta calculation
        B_deg = (n-1)*(360/365)
    
        #Convert B into radians
        B = B_deg*np.pi/180
    
        #Calculate delta (in degrees) - declination - angle of the sun at solar noon i.e. whent the sun is on the local meridian
        delta_deg = (180/(np.pi))*(0.006918-0.399912*np.cos(B)+0.070257*np.sin(B)-0.006758*np.cos(2*B)+0.000907*np.sin(2*B)-0.002697*np.cos(3*B)+0.00148*np.sin(3*B))
    
        #Convert delta into radians
        delta = delta_deg*np.pi/180
    
        #Calculate omega (in degrees) - hour angle
        omega_deg = h*15-187.5
        #Convert omega to radians
        omega = omega_deg*np.pi/180
    
        #Calculate theta (in radians) - Angle of incidence on the collector plate
        theta = np.arccos((np.sin(delta)*np.sin(lat-beta))+(np.cos(delta)*np.cos(omega)*np.cos(lat-beta)))
    
        #Convert theta to degrees for use in angular beam transmition absorption product calculation
        theta_deg = theta*180/np.pi
        
        #Allow function to be exited early if there is no solar radiation incident on the collector
        #if I == 0 and Ibn == 0:
            #Qu = 0
            #It_col = 0
            #Function_outputs = [Qu,It_col,theta_deg]
            #return Function_outputs
    
        #Calculate the beam radiation incident on the tilted collector surface
        #Ibt = Ibn*(np.cos(theta))

        #Calculate the beam radiation incident on a horizontal surface
        #Ibh = Ibn*((np.sin(delta))*(np.sin(lat))+(np.cos(delta))*(np.cos(omega))*(np.cos(lat)))
    
        #Calculate the diffuse radiation incident on a horizontal surface
        Idh = I - Ibh
    
        #Calculate equivalent theta for ground reflected radiation
        theta_tag = 90-(0.5788*beta)+(0.002693*(beta**2))
        #print('theta_tag:',theta_tag)
    
        #Calculate equivalent theta for diffuse radiation
        theta_tad = 59.7-(0.1388*beta)+(0.001497*(beta**2))
        #print('theta_tad:',theta_tad)
    
        #Calculate normal transimition absorption product
        tan = 1.01*t*a
        #print('tan:', tan)
    
        #Calculate the angular transmition absorption product for beam radiation - using IAM specific to TVP collector
        if theta_deg <= 30:
            #If theta is less than 30 then transmission absorption product is just equal to the normal value
            tab = tan
        elif theta_deg > 30 and theta_deg < 90:
            #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
            tab = tan*((-1*(10**(-8))*theta_deg**4)+(-2*(10**(-6))*theta_deg**3)+(0.0002*theta_deg**2)-(0.003*theta_deg)+1.0061)
        else:
            tab = 0
            #print('No beam radiation incident on the collector')
        
        #print('tab:',tab)
    
        #Calculate the angular transmition absorption product for ground reflected radiation
        if theta_tag <= 30:
            #If theta is less than 30 then transmission absorption product is just equal to the normal value
            tag = tan
        elif theta_tag > 30 and theta_tag < 90:
            #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
            tag = tan*((-1*(10**(-8))*theta_tag**4)+(-2*(10**(-6))*theta_tag**3)+(0.0002*theta_tag**2)-(0.003*theta_tag)+1.0061)
        else:
            tag = 0
        
        #print('tag:',tag)
    
        #Calculate the angular transmition absorption product for diffuse radiation
        if theta_tad <= 30:
            #If theta is less than 30 then transmission absorption product is just equal to the normal value
            tad = tan
        elif theta_tad > 30 and theta_tad < 90:
            #If theta is between 30 and 90 degrees transmission absortion product is given by the polynomial
            tad = tan*((-1*(10**(-8))*theta_tad**4)+(-2*(10**(-6))*theta_tad**3)+(0.0002*theta_tad**2)-(0.003*theta_tad)+1.0061)
        else:
            tad = 0
        
        #print('tad:', tad)
    
        #Calculate Rb
        Rb = ((np.cos(lat-beta))*(np.cos(delta))*(np.cos(omega))+(np.sin(lat-beta))*(np.sin(delta)))/((np.cos(lat))*(np.cos(delta))*(np.cos(omega))+(np.sin(lat))*(np.sin(delta)))
        #Rb = 1
    
        #Calculate the absorbed solar radiation per unit collector area
        #S = Ib*Rb*tab + Id*tad*((1+np.cos(beta))/2) + I*Pg*tag*((1-np.cos(beta))/2)
        #print('S:',S)
    
        #Calculate the absorbed solar radiation per unit collector area
        S = Ibh*Rb*tab + Idh*tad*((1+np.cos(beta))/2) + I*Pg*tag*((1-np.cos(beta))/2)
        #print('S:',S)
    
        #Include factor for reduced absorption due to dust
        Df = 0.99

        #Include factor for reduced absoprtion due to self shading
        Ds = 0.99
    
        #Calculated adjusted shading when dust and self shading are taken into account
        Sa = S*Df*Ds
        #print('Sa:', Sa)
    
        #Calaculate initial Cp estimate
        Cp = CP.PropsSI('Cp0mass','T',Tfi+273,'P',P,'water')
    
        #Calculate initial Ul estimate
        Ul = Ulb*(Tfi-Ta)+Ula
    
        #Calculate initial Fr estimate
        Fr = ((m_dot*Cp)/(Ac*Ul))*(1-np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp)))
    
        #Calculate initial estimate of the useful energy gain
        Qu = Ac*Fr*(Sa-Ul*(Tfi-Ta))
    
        #Set limit for the number of iterations that can be carried out
        #Q = 100
        
        #Calculate inlet pipe's heat loss area
        Ai = np.pi*Di*Li
        
        #Calculate outlet pipe's heat loss area
        Ao = np.pi*Do*Lo
        
        #Calculate the losses through the pipe inlet
        Lin = Ud_array*Ai*(Tfi-Ta)
    
        #for r in range(Q):
            #if r==(Q-2):
                #print('Tfm iteration has not converged to the required value',v)
    
        #Define i so that a counter for the loop can be used
        i = 2
        
            #Set a limit for the number of iterations that can be carried out
        Z = 50
    
        for n in range(Z):
        
            #Use counter to inform user on whether or not the iteration has converged to within 1%
            i += 1
            if i == Z:
                print('Qu iteration has not converged to within 1%',v)
                    
            #Calculate the collector inlet temperature
            Tci = Tfi - Lin/(m_dot*Cp)
        
            #Calculate new Tfm value
            Tfm = Tci+(Qu/(Ac*Fr*Ul))*(1-(Fr/F_dash))
            
            #Calculate new Ul value using Tfm
            Ul = Ulb*(Tfm-Ta)+Ula
            
            #Calculate new Cp value using Tfm
            Cp = CP.PropsSI('Cp0mass','T',Tfm+273,'P',P,'water')
        
            #Calculate new Fr value with new Cp value
            Fr = ((m_dot*Cp)/(Ac*Ul))*(1-np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp)))
        
            #Calculate Tpm estimate
            Tpm = Tfi+(Qu/(Ac*Fr*Ul))*(1-Fr)
        
            #Introduce Qu_est so the value of Qu in this iteration can be compared to the value in the last iteration.
            Qu_est = Qu
        
            #Calculate new Qu estimate
            Qu = Ac*(Sa-Ul*(Tpm-Ta))
        
                
            #Terminate loop if the new value of Qu is within 1% of the previous value
            if np.all(np.logical_and((abs(Qu) > 0.99*abs(Qu_est)),(abs(Qu) < 1.01*abs(Qu_est)))):
            #if abs(Qu) > 0.99*abs(Qu_est) and abs(Qu) < 1.01*abs(Qu_est):
                break
            else:
                continue
            
            #Provide if function so that an iteration to the desired mean temperature is carried out
            #if Tfm > (Tfm_tar+1):
                #Tfi -= 0.5
            #elif Tfm < (Tfm_tar-1):
                #Tfi += 0.5
            #else:
                #break
    
        #print('Tfm:',Tfm)
        #print('Cp:',Cp)
        #print('Fr:',Fr)
        #print('Tpm:',Tpm)
        #print('Qu:',Qu)
        
        #Calculate the collector fluid outlet temperature
        Tco = Ta + (Sa/Ul) + (Tfi-Ta-(Sa/Ul))*np.exp(-(Ac*Ul*F_dash)/(m_dot*Cp))
        #print('Tfo',Tfo)
        
        #Calculate an initial estimate of the outlet pipe losses
        Lout = Ud_array*Ao*(Tco-Ta)
        
        for ii in range(200):
            
            if ii == 198:
                print('Lout has not converged to sufficient accuracy')
            
            #Calculate new estimate of the system fluid outlet temperature
            Tfo = Tco - Lout/(m_dot*Cp)
            
            #Introduce Loiut_est so the value of Lout in this iteration can be compared to the value in the last iteration.
            Lout_est = Lout
            
            #Calculate a new estimate of Lout
            Lout = Ud_array*Ao*(Tfo-Ta)
            
            #Terminate loop if the new value of Lout is within 1% of the previous value
            if np.all(np.logical_and((abs(Lout) > 0.99*abs(Lout_est)),(abs(Lout) < 1.01*abs(Lout_est)))):
            #if abs(Lout) > 0.99*abs(Lout_est) and abs(Lout) < 1.01*abs(Lout_est):
                break
            else:
                continue
            
        #Calculate the power output taking into account pipe losses
        Qu_loss = Qu - Lout - Lin
        
        #Set any results of negative power output to zero
        for i in range(len(Qu_loss)):
            if Qu_loss[i] < 0:
                Qu_loss[i]=0

        #Set negative power outputs to zero
        #if Qu_loss < 0:
            #Qu_loss = 0
            
        #Calculate Collector Efficiency
        #eff = Qu/(I*Ag)
        #print('Efficiency:',eff)
    
        #Calculate difference between the mean fluid temperature and ambient temperature
        T_diff = Tfm - Ta
        #print('T_diff:',T_diff)
    
        #Calculate difference between mean plate and mean fluid temperatures
        T_grad = Tpm - Tfm
    
        #Calculate beam contribution to radiation incident on the collector
        Ibt_col = Ac*Ibh*Rb
    
        #Calculate the diffuse contribution to the radiation incident on the collector
        Idt_col = Ac*Idh*((1+np.cos(beta))/2)
    
        #Calculate ground reflected contribution to the radiation incident on the collector
        Igt_col = Ac*I*Pg*((1-np.cos(beta))/2)
    
        #Calculate total radaition incident on the collector
        It_col = Ibt_col+Idt_col+Igt_col
    
        #Calculate the output energy for that hour
        E_out = Qu_loss*3600
        
        #Put function outputs in a list
        #Function_outputs = [Qu,Tfo,eff,T_diff,Ul,T_grad,Tfm,Tpm,It_col,Ibt_col,Idt_col,Igt_col,E_out]
        Function_outputs = [Qu_loss, It_col]
        
        return Function_outputs

    #Output an array of useful power results
    Qu_array = absorbed_radiation(I,Ibh,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Ud_array,Li,Lo,Di,Do)[0]
    
    #Output an array of collector efficiency results
    #eff_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[2]

    #Output an array of the results for the temperature difference between the mean fluid temperature and ambient temperature
    #T_diff_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[3]

    #Output array for Ul results
    #Ul_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[4]

    #Output array of plate to fluid temperature difference results
    #T_grad_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[5]

    #Output array of fluid mean temperature
    #Tfm_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[2]

    #Output array of plate mean temperature
    #Tpm_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[7]

    #Output array of total incident radiation on the plate
    It_col_array = absorbed_radiation(I,Ibh,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb,Ud_array,Li,Lo,Di,Do)[1]

    #Output array of total beam radiation on the plate
    #Ibt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[9]

    #Output array of total diffuse radiation on the plate
    #Idt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[10]

    #Output array of total diffuse radiation on the plate
    #Igt_col_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[11]

    #Output array for the energy output for that hour
    #E_out_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[12]
    
    #Output array for angle of incidence for that hour
    #theta_deg_array = absorbed_radiation(I,Ibn,Pg,beta,lat,n,h,t,a,Ac,Ag,Ta,m_dot,F_dash,P,Tfi,Ula,Ulb)[2]
    
    #Give each hourly energy output in kilowatt hours
    Qu_kW = Qu_array/1000
    
    #Give each hourly incident horizonatal radiation value in kilowatt hours
    I_kW = (I*Ac)/1000
    
    #Give each hourly incident radiation value in kilowatt hours
    It_col_kW = It_col_array/1000
    
    #Append the hourly energy output to the energy output list
    Qu_list.append(Qu_kW)
    
    Qu_tot += Qu_kW
    
    It_col_sum += It_col_kW
    
    #I_sum += I_kW
    
    #for u in range(len(np.size(Qu_kW))):
        #Qu_list.append(Qu_kW[u])
        #It_col_list.append(It_col_kW[u])
    
    #Tfm_sum += Tfm_array
    
    #T_diff_sum += T_diff_array
    
    #print(Tfm_sum)
    
    #if n == 31 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0 
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 59 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(28*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(28*24))
        #T_diff_sum = 0
    #elif n == 90 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 120 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(30*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(30*24))
        #T_diff_sum = 0
    #elif n == 151 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 181 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(30*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(30*24))
        #T_diff_sum = 0
    #elif n == 212 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 243 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 273 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(30*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(30*24))
        #T_diff_sum = 0
    #elif n == 304 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #elif n == 334 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(30*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(30*24))
        #T_diff_sum = 0
    #elif n == 365 and h == 24:
        #Qu_tot_list.append(Qu_tot)
        #Qu_tot = 0
        #It_tot_list.append(It_col_sum)
        #It_col_sum = 0
        #I_list.append(I_sum)
        #I_sum = 0
        #Tfm_list.append(Tfm_sum/(31*24))
        #Tfm_sum = 0
        #T_diff_list.append(T_diff_sum/(31*24))
        #T_diff_sum = 0
    #else:
        #continue
        
    
    #Append hourly incident radiation to incident energy list
    #It_col_list.append(It_col_array)
    
    #Append hourly incident value to incidence list
    #theta_deg_list.append(theta_deg_array)
    
    #Calculate total power output for time period in question
    #Qu_tot_array += Qu_kW
    
#Define a variable to add all the individual hourly values in Qu_list
#Qu_all = 0

#Use for loop to calculate the total energy value of iteration
#for a in range(len(Qu_list)):
    #Qu_all += Qu_list[a]

#Qu_year = 0
    
#for b in range(len(Qu_tot_list)):
    #Qu_year += Qu_tot_list[b]

#print(theta_deg_list)
#print(It_col_list)
#print(Qu_tot_list)    
#print(Qu_all)
#print(Qu_tot_list)
#print(Qu_year)
print(It_tot_list)
print(Tfm_list)
print(T_diff_list)
print(I_list)
print(Qu_tot)
print(It_col_sum)
    
    #print('Tfm:',Tfm_array)
    #print('Qu:',Qu_array)
    #print('It:',It_col_array)
    #print('Ibt:',Ibt_col_array)
    #print('Idt:',Idt_col_array)
    #print('Igt:',Igt_col_array)
    #print('E_out:',E_out_array)

#print(absorbed_radiation(Ib,Id,Pg,beta,theta_deg,t,a,Ac,Ul,Tfi,Ta,m_dot,F_dash,P))
    
#PL.plot(T_diff_array,eff_array)
#PL.plot(T_diff_array,Ul_array)

#Create a workbook to write to
#wb = Workbook()
wb = load_workbook("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Monthly_collector_output_data.xlsx")

#Create a sheet to write to
#sheet1 = wb.add_sheet('Sheet 2')

#for i in range(len(eff_array)):
    
    #Write temperature difference data to the sheet
    #sheet1.write(i,2,T_diff_array[i])
    
    #Write efficiency data to the sheet
    #sheet1.write(i, 3, eff_array[i])

#Save the spreadsheet
#wb.save("Fifth attempt at writing to a spreadsheet.xls")

#Define which worksheet to write data to
ws = wb["Effect_of_Losses"] 

for i in range(len(Qu_tot)):

    #Define cell to write data to
    wcell1 = ws.cell(i+3,37)
    wcell2 = ws.cell(i+3,36)
    wcell3 = ws.cell(i+3,35)
    #wcell4 = ws.cell(i+3,48)
    #wcell5 = ws.cell(i+2,53)

    #State what value should be written to the cell
    wcell1.value = Qu_tot[i]
    wcell2.value = It_col_sum
    wcell3.value = Ud_list[i]
    #wcell4.value = T_diff_list[i]
    #wcell5.value = Tpm_array[i]

#Save the spreadsheet
wb.save("C:\\Users\\maxaj\\OneDrive\\Thermoacoustic project\\Python Code\\Monthly_collector_output_data.xlsx")